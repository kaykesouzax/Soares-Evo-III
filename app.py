from flask import Flask, request, send_file, render_template, jsonify, send_from_directory
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import io
import os
import re

app = Flask(__name__)

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'), 'favicon.ico', mimetype='image/x-icon')

# ============================================================
# BASE MESTRA - RANKING
# ============================================================
MAPA_NOMES = {
    "ALEXSANDRO ALVES": "ALEXSANDRO SOUZA", "ALEXSANDRO SOUZA": "ALEXSANDRO SOUZA", "ALEXSANDRO ALVES DE SOUZA": "ALEXSANDRO SOUZA",
    "ANGELA ALMEIDA": "ANGELA ALMEIDA", "ANGELA SILVA": "ANGELA ALMEIDA", "ANGELA ALMEIDA DA SILVA": "ANGELA ALMEIDA",
    "DOUGLAS MACIEL": "DOUGLAS MACIEL", "DOUGLAS DIAS": "DOUGLAS MACIEL", "DOUGLAS MACIEL DIAS": "DOUGLAS MACIEL",
    "TAYANE MAIA": "TAYANE MAIA", "TAYANE SILVA": "TAYANE MAIA", "TAYANE MAIA E SILVA": "TAYANE MAIA",
    "ARTEMISA SOUZA": "ARTEMISA BELEM", "ARTEMISA BELEM": "ARTEMISA BELEM", "ARTEMISA BELEM DE SOUZA": "ARTEMISA BELEM",
    "ANTONIA AGUIDA": "AGUIDA SILVA", "AGUIDA NASCIMENTO": "AGUIDA SILVA", "AGUIDA SILVA": "AGUIDA SILVA", "ANTONIA AGUIDA NASCIMENTO DA SILVA": "AGUIDA SILVA",
    "RODRIGO QUEMEL": "RODRIGO QUEMEL", "RODRIGO SANTOS": "RODRIGO QUEMEL", "RODRIGO DOS SANTOS QUEMEL": "RODRIGO QUEMEL",
    "JANDERSON CARNEIRO": "JANDERSON CARNEIRO", "JANDERSON GUSTAVO": "JANDERSON CARNEIRO", "JANDERSON GUSTAVO CARNEIRO": "JANDERSON CARNEIRO",
    "JOSE EWERTHON": "JOSE EWERTHON", "EWERTHON BARROS": "JOSE EWERTHON", "JOSE EWERTHON BARROS CAVALCANTE": "JOSE EWERTHON",
    "THALYS CASTRO": "THALYS CASTRO", "THALYS SILVA": "THALYS CASTRO", "THALYS CASTRO DA SILVA": "THALYS CASTRO",
    "ANTONIO MARCOS": "ANTONIO MARCIO", "ANTONIO BATISTA": "ANTONIO MARCIO", "ANTONIO MARCOS BATISTA DOS SANTOS": "ANTONIO MARCIO",
    "ANTONIO MARCIO BATISTA DOS SANTOS": "ANTONIO MARCIO", "ANTONIO MARCIO": "ANTONIO MARCIO",
    "GUILHERME DO NASCIMENTO MELO": "GUILHERME MELO", "GUILHERME DO": "GUILHERME MELO", "GUILHERME NASCIMENTO": "GUILHERME MELO", "GUILHERME MELO": "GUILHERME MELO",
    "ANTONIO HENRIQUE": "ANTONIO HENRIQUE", "ANTONIO COSTA": "ANTONIO HENRIQUE", "ANTONIO HENRIQUE DA COSTA DOS SANTOS": "ANTONIO HENRIQUE",
    "OMAR VALE": "OMAR VALE", "OMAR NASCIMENTO": "OMAR VALE", "OMAR VALE NASCIMENTO": "OMAR VALE",
    "WENDEL SILVA": "WENDEL SILVA", "WENDEL OLIVEIRA": "WENDEL SILVA", "WENDEL SILVA DE OLIVEIRA": "WENDEL SILVA",
    "BRUNA CARLA BEZERRA": "BRUNA CARLA", "BRUNA BEZERRA": "BRUNA CARLA", "BRUNA CARLA BEZERRA SOUZA": "BRUNA CARLA", "BRUNA SOUZA": "BRUNA CARLA",
    "JONAS DE BARROS": "JONAS COELHO", "JONAS BARROS": "JONAS COELHO", "JONAS COELHO": "JONAS COELHO", "JONAS DE BARROS COELHO": "JONAS COELHO", "JONAS DE": "JONAS COELHO",
    "CARLOS BARBOSA": "CARLOS BARBOSA", "CARLOS SOUZA": "CARLOS BARBOSA", "CARLOS BARBOSA DE SOUZA": "CARLOS BARBOSA",
    "LUAN HENRIQUE": "LUAN HENRIQUE", "LUAN DIAS": "LUAN HENRIQUE", "LUAN HENRIQUE ROCHA DIAS": "LUAN HENRIQUE",
    "ANDREI ALVES": "ANDREI ALVES", "ANDREI SOUZA": "ANDREI ALVES", "ANDREI ALVES DE SOUZA": "ANDREI ALVES",
    "MARIA HELENA": "MARIA HELENA", "MARIA GOMES": "MARIA HELENA", "MARIA HELENA PIMENTA GOMES": "MARIA HELENA",
    "ADRIANO JOTAERRY": "ADRIANO JOTAERRY", "ADRIANO NUNES": "ADRIANO JOTAERRY", "ADRIANO JOTAERRY VIEIRA NUNES": "ADRIANO JOTAERRY",
    "SERGIO MOREIRA": "SERGIO MOREIRA", "SERGIO COSTA": "SERGIO MOREIRA", "SERGIO MOREIRA DA COSTA JUNIOR": "SERGIO MOREIRA",
    "EDSON RUAN": "EDSON RUAN", "EDSON NOGUEIRA": "EDSON RUAN", "EDSON RUAN LEAL NOGUEIRA": "EDSON RUAN",
    "KAWANE ANDRADE": "KAWANE KISTNER", "KAWANE KISTNER": "KAWANE KISTNER", "KAWANE ANDRADE KISTNER": "KAWANE KISTNER",
    "ROSIELE TORRES": "ROSIELE TORRES", "ROSIELE SILVA": "ROSIELE TORRES", "ROSIELE DA SILVA TORRES": "ROSIELE TORRES",
    "SHIRLANE MELO": "SHIRLANE MELO", "SHIRLANE SANTANA": "SHIRLANE MELO", "SHIRLANE SANTANA DE MELO": "SHIRLANE MELO",
    "JOSE WILLIAN": "JOSE WILLIAN", "JOSE PINTO": "JOSE WILLIAN", "JOSE WILLIAN SILVA PINTO": "JOSE WILLIAN",
    "JOSE SILVA": "JOSE WILLIAN", "JOSE WILLIAM": "JOSE WILLIAN",
    "DIEGO PEREIRA": "DIEGO GOMES", "DIEGO GOMES": "DIEGO GOMES", "DIEGO PEREIRA GOMES": "DIEGO GOMES",
    "RAIMISON RODRIGUES": "RAIMISON RODRIGUES", "RAIMISON FRANCA": "RAIMISON RODRIGUES",
    "RAIMISON DE FRANCA RODRIGUES": "RAIMISON RODRIGUES", "RAIMISON DE": "RAIMISON RODRIGUES", "RAIMISON DE FRANCA": "RAIMISON RODRIGUES",
}

VENDEDORES_PDV = {
    "ALEXSANDRO SOUZA": "LÁBREA", "ANGELA ALMEIDA": "LÁBREA", "DOUGLAS MACIEL": "LÁBREA",
    "TAYANE MAIA": "LÁBREA", "ARTEMISA BELEM": "LÁBREA", "AGUIDA SILVA": "LÁBREA",
    "RODRIGO QUEMEL": "LÁBREA", "JANDERSON CARNEIRO": "LÁBREA", "JOSE EWERTHON": "LÁBREA",
    "THALYS CASTRO": "LÁBREA", "ANTONIO MARCIO": "LÁBREA", "GUILHERME MELO": "LÁBREA",
    "ANTONIO HENRIQUE": "BOCA DO ACRE", "OMAR VALE": "BOCA DO ACRE",
    "WENDEL SILVA": "BOCA DO ACRE", "BRUNA CARLA": "BOCA DO ACRE", "JONAS COELHO": "BOCA DO ACRE",
    "CARLOS BARBOSA": "BOCA DO ACRE",
    "LUAN HENRIQUE": "HUMAITÁ", "ANDREI ALVES": "HUMAITÁ", "MARIA HELENA": "HUMAITÁ",
    "ADRIANO JOTAERRY": "HUMAITÁ", "SERGIO MOREIRA": "HUMAITÁ",
    "EDSON RUAN": "HUMAITÁ", "KAWANE KISTNER": "HUMAITÁ", "ROSIELE TORRES": "HUMAITÁ",
    "SHIRLANE MELO": "HUMAITÁ", "JOSE WILLIAN": "HUMAITÁ", "DIEGO GOMES": "HUMAITÁ",
    "RAIMISON RODRIGUES": "HUMAITÁ",
}

# ============================================================
# LÓGICA - RANKING
# ============================================================
def parse_base(texto):
    resultado = {}
    for linha in texto.strip().splitlines():
        linha = linha.strip()
        if not linha:
            continue
        partes = linha.rsplit(None, 1)
        if len(partes) == 2:
            nome_raw = partes[0].strip().upper()
            try:
                qtd = int(partes[1].strip())
            except ValueError:
                continue
            nome_pad = MAPA_NOMES.get(nome_raw)
            if nome_pad:
                resultado[nome_pad] = resultado.get(nome_pad, 0) + qtd
    return resultado

def ranking(lista):
    s = sorted(lista, key=lambda x: (-x["total"], -x["cotas"], x["nome"]))
    return [{**d, "posicao": i + 1} for i, d in enumerate(s)]

def processar_dados(texto_cotas, texto_novos):
    cotas = parse_base(texto_cotas)
    novos = parse_base(texto_novos)
    todos = set(cotas) | set(novos)
    dados = []
    for nome in todos:
        c = cotas.get(nome, 0)
        n = novos.get(nome, 0)
        t = c + n
        if t == 0:
            continue
        dados.append({"nome": nome, "pdv": VENDEDORES_PDV.get(nome, ""), "cotas": c, "novos": n, "total": t})
    return dados

def get_zerados(dados_ativos):
    ativos = {d["nome"] for d in dados_ativos}
    zerados = sorted([
        {"nome": nome, "pdv": pdv}
        for nome, pdv in VENDEDORES_PDV.items()
        if nome not in ativos
    ], key=lambda x: x["nome"])
    return zerados

# ============================================================
# LÓGICA - ESTOQUE
# ============================================================
def eh_motor(modelo):
    modelo_up = modelo.upper().strip()
    return modelo_up.startswith("GX") or modelo_up == "WHC10XR"

def normalizar_modelo(modelo):
    modelo_up = modelo.upper().strip()
    # GX 160 QX ou GX 160 QD → GX 160
    if modelo_up.startswith("GX 160 Q"):
        return "GX 160"
    return modelo

def padronizar_cor(cor):
    cor_up = cor.upper().strip()
    # Remove PER / PER. / PÉROLA
    cor_up = re.sub(r'\bPER\.?\b', '', cor_up)
    cor_up = re.sub(r'\bPÉROLA\b', '', cor_up)
    # Remove MET / MET. / METALICO / METÁLICO
    cor_up = re.sub(r'\bMET\.?\b', '', cor_up)
    cor_up = re.sub(r'\bMETALICO\b', '', cor_up)
    cor_up = re.sub(r'\bMETÁLICO\b', '', cor_up)
    # Remove espaços extras e pontos finais
    cor_up = re.sub(r'\s+', ' ', cor_up).strip().strip('.')
    return cor_up

def processar_estoque_texto(texto):
    linhas = texto.strip().splitlines()
    dados = {}
    ordem = []
    
    for linha in linhas:
        linha = linha.strip()
        if not linha:
            continue
        partes = linha.split('\t')
        
        if len(partes) == 2:
            modelo, cor = partes
        elif len(partes) >= 3:
            modelo, chassi, cor = partes[0], partes[1], partes[2]
        else:
            continue
        
        modelo = normalizar_modelo(modelo.strip())
        cor_padrao = padronizar_cor(cor.strip())
        
        if modelo not in dados:
            dados[modelo] = set()
            ordem.append(modelo)
        
        if not eh_motor(modelo) and cor_padrao:
            dados[modelo].add(cor_padrao)
    
    return dados, ordem

def processar_estoque_excel(arquivo):
    # Salvar temporariamente
    import tempfile
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        arquivo.save(tmp.name)
        tmp_path = tmp.name
    
    try:
        # Tentar ler ignorando warnings de estilos
        import warnings
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        
        wb = load_workbook(tmp_path, data_only=True, read_only=False, keep_links=False)
        ws = wb.active
        
        dados = {}
        ordem = []
        
        # Iterar pelas linhas manualmente
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            
            try:
                modelo = str(row[0]).strip()
                cor = str(row[2]).strip() if len(row) >= 3 and row[2] else ""
            except:
                continue
            
            if not modelo or modelo == 'None':
                continue
            
            modelo = normalizar_modelo(modelo)
            cor_padrao = padronizar_cor(cor)
            
            if modelo not in dados:
                dados[modelo] = set()
                ordem.append(modelo)
            
            if not eh_motor(modelo) and cor_padrao:
                dados[modelo].add(cor_padrao)
        
        wb.close()
        
    except Exception as e:
        # Se falhar, tentar abordagem alternativa lendo como binário
        import openpyxl
        from io import BytesIO
        
        arquivo.seek(0)
        wb = openpyxl.load_workbook(BytesIO(arquivo.read()), data_only=True)
        ws = wb.active
        
        dados = {}
        ordem = []
        
        for row in ws.values:
            if not row or not row[0]:
                continue
            
            try:
                modelo = str(row[0]).strip()
                cor = str(row[2]).strip() if len(row) >= 3 and row[2] else ""
            except:
                continue
            
            if not modelo or modelo == 'None':
                continue
            
            modelo = normalizar_modelo(modelo)
            cor_padrao = padronizar_cor(cor)
            
            if modelo not in dados:
                dados[modelo] = set()
                ordem.append(modelo)
            
            if not eh_motor(modelo) and cor_padrao:
                dados[modelo].add(cor_padrao)
        
        wb.close()
    finally:
        import os
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    
    return dados, ordem

def consolidar_estoque(dados, ordem):
    normais = []
    motores = []
    
    for modelo in ordem:
        cores_set = dados.get(modelo, set())
        cores_lista = sorted(cores_set)
        cores_str = " – ".join(cores_lista) if cores_lista else ""
        
        if eh_motor(modelo):
            motores.append({"modelo": modelo, "cor": ""})
        else:
            normais.append({"modelo": modelo, "cor": cores_str})
    
    return normais + motores

# ============================================================
# GERAÇÃO EXCEL - RANKING
# ============================================================
VERDE_ESCURO = "006400"
VERDE_CLARO  = "90EE90"
BRANCO       = "FFFFFF"

def borda():
    s = Side(style="thin", color=VERDE_CLARO)
    return Border(left=s, right=s, top=s, bottom=s)

def preencher_aba(ws, dados):
    cab = ["Posição", "Nome do Vendedor", "Cotas", "Novos", "Total Geral"]
    for ci, txt in enumerate(cab, 1):
        c = ws.cell(row=1, column=ci, value=txt)
        c.font      = Font(name="Calibri", bold=True, color=BRANCO, size=12)
        c.fill      = PatternFill("solid", start_color=VERDE_ESCURO)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borda()
    ws.row_dimensions[1].height = 20

    for ri, item in enumerate(dados, 2):
        ws.row_dimensions[ri].height = 15
        for ci, v in enumerate([item["posicao"], item["nome"], item["cotas"], item["novos"], item["total"]], 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font      = Font(name="Calibri", size=11)
            c.fill      = PatternFill("solid", start_color=BRANCO)
            c.alignment = Alignment(horizontal="left" if ci == 2 else "center", vertical="center")
            c.border    = borda()

    tr = len(dados) + 2
    ws.row_dimensions[tr].height = 15
    soma_c = sum(d["cotas"] for d in dados)
    soma_n = sum(d["novos"] for d in dados)
    soma_t = sum(d["total"] for d in dados)
    for ci, v in enumerate(["", "Total Geral", soma_c, soma_n, soma_t], 1):
        c = ws.cell(row=tr, column=ci, value=v)
        c.font      = Font(name="Calibri", bold=True, size=11)
        c.fill      = PatternFill("solid", start_color=BRANCO)
        c.alignment = Alignment(horizontal="left" if ci == 2 else "center", vertical="center")
        c.border    = borda()

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

def preencher_aba_zerados(ws, zerados):
    cab = ["Posição", "Nome do Vendedor", "PDV", "Cotas", "Novos", "Total Geral"]
    for ci, txt in enumerate(cab, 1):
        c = ws.cell(row=1, column=ci, value=txt)
        c.font      = Font(name="Calibri", bold=True, color=BRANCO, size=12)
        c.fill      = PatternFill("solid", start_color=VERDE_ESCURO)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borda()
    ws.row_dimensions[1].height = 20

    for ri, item in enumerate(zerados, 2):
        ws.row_dimensions[ri].height = 15
        vals = [ri - 1, item["nome"], item["pdv"], 0, 0, 0]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font      = Font(name="Calibri", size=11)
            c.fill      = PatternFill("solid", start_color=BRANCO)
            c.alignment = Alignment(horizontal="left" if ci == 2 else "center", vertical="center")
            c.border    = borda()

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14

def gerar_excel_ranking(dados_todos):
    wb = Workbook()

    geral   = ranking(dados_todos)
    labrea  = ranking([d for d in dados_todos if d["pdv"] == "LÁBREA"])
    boca    = ranking([d for d in dados_todos if d["pdv"] == "BOCA DO ACRE"])
    humaita = ranking([d for d in dados_todos if d["pdv"] == "HUMAITÁ"])
    zerados = get_zerados(dados_todos)

    ws = wb.active; ws.title = "GERAL";        preencher_aba(ws, geral)
    ws = wb.create_sheet("LÁBREA");             preencher_aba(ws, labrea)
    ws = wb.create_sheet("BOCA DO ACRE");       preencher_aba(ws, boca)
    ws = wb.create_sheet("HUMAITÁ");            preencher_aba(ws, humaita)
    ws = wb.create_sheet("ZERADOS");            preencher_aba_zerados(ws, zerados)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ============================================================
# GERAÇÃO EXCEL - ESTOQUE
# ============================================================
def gerar_excel_estoque(dados_consolidados):
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque"
    
    # Linha 1 - Título
    ws.merge_cells('A1:B1')
    c = ws['A1']
    c.value = "ESTOQUE DISPONÍVEL"
    c.font = Font(name="Calibri", bold=True, size=12)
    c.fill = PatternFill("solid", start_color="999999")
    c.alignment = Alignment(horizontal="center", vertical="center")
    
    # Linha 2 - Cabeçalho
    for ci, txt in enumerate(["MODELO", "COR"], 1):
        c = ws.cell(row=2, column=ci, value=txt)
        c.font = Font(name="Calibri", bold=True, size=11)
        c.fill = PatternFill("solid", start_color="DDDDDD")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    # Dados
    for ri, item in enumerate(dados_consolidados, 3):
        for ci, v in enumerate([item["modelo"], item["cor"]], 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Calibri", size=11)
            c.alignment = Alignment(horizontal="left" if ci == 1 else "left", vertical="center")
            c.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
    
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ============================================================
# ROTAS
# ============================================================
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/preview", methods=["POST"])
def preview():
    data = request.get_json()
    texto_cotas = data.get("cotas", "")
    texto_novos = data.get("novos", "")
    dados = processar_dados(texto_cotas, texto_novos)
    if not dados:
        return jsonify({"erro": "Nenhum vendedor identificado."}), 400

    def rank(lista):
        s = sorted(lista, key=lambda x: (-x["total"], -x["cotas"], x["nome"]))
        return [[i+1, d["nome"], d["pdv"], d["cotas"], d["novos"], d["total"]] for i, d in enumerate(s)]

    zerados = get_zerados(dados)

    return jsonify({
        "geral":   rank(dados),
        "labrea":  rank([d for d in dados if d["pdv"] == "LÁBREA"]),
        "boca":    rank([d for d in dados if d["pdv"] == "BOCA DO ACRE"]),
        "humaita": rank([d for d in dados if d["pdv"] == "HUMAITÁ"]),
        "zerados": [[i+1, z["nome"], z["pdv"], 0, 0, 0] for i, z in enumerate(zerados)],
    })

@app.route("/gerar", methods=["POST"])
def gerar():
    data = request.get_json()
    texto_cotas = data.get("cotas", "")
    texto_novos = data.get("novos", "")

    dados = processar_dados(texto_cotas, texto_novos)
    if not dados:
        return jsonify({"erro": "Nenhum vendedor identificado. Verifique os dados colados."}), 400

    buf = gerar_excel_ranking(dados)
    nome_arquivo = datetime.now().strftime("%d.%m") + ".xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=nome_arquivo
    )

@app.route("/estoque", methods=["POST"])
def estoque():
    try:
        # Prioridade: arquivo Excel > texto
        if 'arquivo' in request.files and request.files['arquivo'].filename:
            arquivo = request.files['arquivo']
            try:
                dados, ordem = processar_estoque_excel(arquivo)
            except Exception as e:
                return jsonify({"erro": f"Erro ao processar Excel: {str(e)}"}), 400
        else:
            texto = request.form.get('texto', '')
            if not texto.strip():
                return jsonify({"erro": "Nenhum dado fornecido."}), 400
            dados, ordem = processar_estoque_texto(texto)
        
        consolidado = consolidar_estoque(dados, ordem)
        buf = gerar_excel_estoque(consolidado)
        
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="Estoque Atual.xlsx"
        )
    except Exception as e:
        return jsonify({"erro": f"Erro interno: {str(e)}"}), 500

if __name__ == "__main__":
    app.run(debug=False)
